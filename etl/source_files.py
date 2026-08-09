"""Discovery of source files and conversion of .xlsx to .csv.

The database loads CSV only, so any .xlsx handed to the framework is converted
first. The converted file keeps the same stem and lands beside the original.
"""

import csv
import datetime as dt
from decimal import Decimal
from pathlib import Path
from typing import List, Tuple

import openpyxl

SUPPORTED_SUFFIXES = {".csv", ".xlsx"}


def list_source_files(source_dir: Path) -> List[Path]:
    """Return supported source files, one entry per logical file.

    Converting an .xlsx leaves a .csv of the same stem beside it. Listing both
    would grow the menu on every run and present one logical file twice, so a
    .csv that shadows an .xlsx of the same stem is treated as the derived
    artefact it is and hidden. The .xlsx remains the thing to select.
    """
    if not source_dir.exists():
        return []

    candidates = [
        path
        for path in source_dir.iterdir()
        if path.is_file() and path.suffix.lower() in SUPPORTED_SUFFIXES
    ]
    xlsx_stems = {
        path.stem for path in candidates if path.suffix.lower() == ".xlsx"
    }
    visible = [
        path
        for path in candidates
        if not (path.suffix.lower() == ".csv" and path.stem in xlsx_stems)
    ]
    return sorted(visible, key=lambda p: p.name.lower())


def _format_cell(value) -> str:
    """Render a spreadsheet cell as the text that will land in the CSV.

    Dates become ISO-8601 so they cast cleanly later. Numbers that are whole
    are written without a trailing .0, which otherwise blocks an integer cast.
    """
    if value is None:
        return ""
    if isinstance(value, dt.datetime):
        if value.time() == dt.time(0, 0):
            return value.date().isoformat()
        return value.isoformat(sep=" ")
    if isinstance(value, dt.date):
        return value.isoformat()
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    if isinstance(value, float):
        if value.is_integer():
            return str(int(value))
        return repr(value)
    if isinstance(value, Decimal):
        return str(value)
    return str(value).strip()


def convert_xlsx_to_csv(xlsx_path: Path, sheet_name: str = None) -> Tuple[Path, int]:
    """Convert an .xlsx file to .csv, retaining the same file stem.

    Returns (csv_path, data_row_count). Trailing fully-empty rows produced by
    spreadsheet editors are discarded.
    """
    workbook = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)
    try:
        worksheet = workbook[sheet_name] if sheet_name else workbook.worksheets[0]
        rows = [
            [_format_cell(cell) for cell in row]
            for row in worksheet.iter_rows(values_only=True)
        ]
    finally:
        workbook.close()

    while rows and not any(cell for cell in rows[-1]):
        rows.pop()

    if not rows:
        raise ValueError("{} contains no rows.".format(xlsx_path.name))

    csv_path = xlsx_path.with_suffix(".csv")
    with csv_path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.writer(handle)
        writer.writerows(rows)

    return csv_path, max(len(rows) - 1, 0)


def sheet_names(xlsx_path: Path) -> List[str]:
    workbook = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)
    try:
        return list(workbook.sheetnames)
    finally:
        workbook.close()


def read_csv_header(csv_path: Path) -> List[str]:
    with csv_path.open("r", newline="", encoding="utf-8-sig") as handle:
        reader = csv.reader(handle)
        for row in reader:
            if any(cell.strip() for cell in row):
                return row
    raise ValueError("{} contains no header row.".format(csv_path.name))


def count_csv_data_rows(csv_path: Path) -> int:
    """Count data rows, excluding the header and any fully blank lines."""
    with csv_path.open("r", newline="", encoding="utf-8-sig") as handle:
        reader = csv.reader(handle)
        total = 0
        header_seen = False
        for row in reader:
            if not any(cell.strip() for cell in row):
                continue
            if not header_seen:
                header_seen = True
                continue
            total += 1
    return total
