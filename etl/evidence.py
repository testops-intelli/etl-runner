"""Writing a validation result out as an xlsx evidence workbook.

Console output is read once and scrolls away. A workbook is the artefact that
survives the run: it can be attached to a migration sign-off, diffed against
the previous rehearsal, or handed to somebody who was not at the keyboard.

Formatting follows the same conventions as the other evidence workbooks in
this portfolio: Arial throughout, frozen header row, native date and numeric
types rather than pre-formatted strings, and one concern per sheet.
"""

import datetime as dt
from pathlib import Path
from typing import Dict, List

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

FONT_NAME = "Arial"
HEADER_FILL = PatternFill("solid", fgColor="D9D9D9")
PASS_FONT = Font(name=FONT_NAME, bold=True, color="006100")
FAIL_FONT = Font(name=FONT_NAME, bold=True, color="9C0006")


def _write_sheet(worksheet, headers: List[str], rows: List[List],
                 empty_note: str = None) -> None:
    worksheet.append(headers)
    for index in range(1, len(headers) + 1):
        cell = worksheet.cell(row=1, column=index)
        cell.font = Font(name=FONT_NAME, bold=True)
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(vertical="center")

    if rows:
        for row in rows:
            worksheet.append(row)
    elif empty_note:
        worksheet.append([empty_note])

    for index in range(1, len(headers) + 1):
        longest = len(str(headers[index - 1]))
        for row in rows:
            if index - 1 < len(row):
                longest = max(longest, len(str(row[index - 1])))
        worksheet.column_dimensions[get_column_letter(index)].width = min(
            max(longest + 2, 10), 60)

    for row in worksheet.iter_rows(min_row=2):
        for cell in row:
            if cell.font is None or cell.font.name != FONT_NAME:
                cell.font = Font(name=FONT_NAME)

    worksheet.freeze_panes = "A2"


def write_workbook(path: Path, *, mapping: Dict, status: str,
                   started_at: dt.datetime, comparison: Dict,
                   derivation_errors: List[Dict], auto_aggregates: List[Dict],
                   declared_aggregates: List[Dict], coverage: List[Dict],
                   key_columns: List[str], source_rows: int,
                   target_rows: int) -> Path:
    workbook = Workbook()

    summary = workbook.active
    summary.title = "Summary"
    _write_sheet(summary, ["Item", "Value"], [
        ["Mapping", mapping["mapping_name"]],
        ["Source", "{}.{}".format(
            mapping["source_schema"], mapping["source_table"])],
        ["Target", "{}.{}".format(
            mapping["target_schema"], mapping["target_table"])],
        ["Row filter", mapping["row_filter"] or "(none)"],
        ["Identifying key", " + ".join(key_columns)],
        ["Validated at", started_at.replace(microsecond=0)],
        ["Result", status],
        ["Source rows", source_rows],
        ["Target rows", target_rows],
        ["Rows matched", comparison["matched"]],
        ["Rows missing from target", len(comparison["missing"])],
        ["Rows in target with no source", len(comparison["orphans"])],
        ["Duplicate keys in target", len(comparison["duplicates"])],
        ["Value mismatches", len(comparison["mismatches"])],
        ["Rows that could not be derived", len(derivation_errors)],
        ["Aggregate checks run",
         len(auto_aggregates) + len(declared_aggregates)],
        ["Aggregate checks failed", sum(
            1 for check in auto_aggregates + declared_aggregates
            if check["status"] == "FAIL")],
    ])
    result_cell = summary.cell(row=8, column=2)
    result_cell.font = PASS_FONT if status == "PASS" else FAIL_FONT

    _write_sheet(
        workbook.create_sheet("Value mismatches"),
        ["Source row", "Identifying key", "Column", "Expected", "Actual"],
        [[item["source_row"], item["key"], item["column"],
          item["expected"], item["actual"]]
         for item in comparison["mismatches"]],
        "No value mismatches.",
    )

    _write_sheet(
        workbook.create_sheet("Missing rows"),
        ["Source row", "Identifying key"],
        [[item["source_row"], item["key"]]
         for item in comparison["missing"]],
        "No rows missing from the target.",
    )

    _write_sheet(
        workbook.create_sheet("Orphan rows"),
        ["Identifying key"],
        [[item["key"]] for item in comparison["orphans"]],
        "No target rows without a source row.",
    )

    _write_sheet(
        workbook.create_sheet("Duplicate keys"),
        ["Identifying key", "Expected count", "Actual count"],
        [[item["key"], item["expected_count"], item["actual_count"]]
         for item in comparison["duplicates"]],
        "No duplicate keys in the target.",
    )

    _write_sheet(
        workbook.create_sheet("Aggregates"),
        ["Check", "Kind", "Left", "Left value", "Right", "Right value",
         "Variance", "Tolerance", "Result"],
        [[check["label"],
          kind,
          check["left_label"],
          check["left_value"],
          check["right_label"],
          check["right_value"],
          check["variance"],
          check["tolerance"],
          check["status"]]
         for kind, group in (("derived", auto_aggregates),
                             ("declared", declared_aggregates))
         for check in group],
        "No aggregate checks configured.",
    )

    _write_sheet(
        workbook.create_sheet("Column coverage"),
        ["Target column", "Type", "Mapping rule", "Validation"],
        [[item["column"], item["data_type"], item["treatment"],
          item["validation"]] for item in coverage],
    )

    _write_sheet(
        workbook.create_sheet("Derivation errors"),
        ["Source row", "Column", "Reason"],
        [[item["row"], item["column"], item["reason"]]
         for item in derivation_errors],
        "Every source row produced a target row.",
    )

    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)
    return path
